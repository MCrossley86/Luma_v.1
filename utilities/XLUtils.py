# Import the necessary modules
import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file,sheetName):
    # Load the workbook, select the specified sheet and return max rows
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def get_column_count(file,sheetName):
    # Load the workbook, select the specified sheet and return max columns
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def read_data(file,sheetName,rownum,columnno):
    # Load the workbook, select the specified sheet and return the value of the cell
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def write_data(file,sheetName,rownum,columnno,data):
    # Load the workbook, select the specified sheet, set the value of the cell with the given data and save the changes to the workbook
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)